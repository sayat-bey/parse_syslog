import yaml
import csv
import time
import queue
import re
from threading import Thread
from pprint import pformat, pprint
from openpyxl import load_workbook, Workbook, styles
from sys import argv
from datetime import datetime, timedelta
from pathlib import Path
from netmiko import ConnectHandler
from netmiko.ssh_exception import NetMikoTimeoutException, SSHException

# решение проблемы при подключении к IOS XR
import logging
logging.getLogger('paramiko.transport').disabled = True 


#######################################################################################
# ------------------------------ classes part ----------------------------------------#
#######################################################################################


class CellSiteGateway:
    def __init__(self, ip, host):
        self.hostname = host
        self.ip_address = ip
        self.ssh_conn = None
        self.os_type = "cisco_ios"

        self.connection_status = True       # failed connection status, False if connection fails
        self.connection_error_msg = None    # connection error message

        self.show_isis_log = None
        self.show_mac_log = None
        self.show_arp_log = None
        self.show_description_log = None
        self.show_isis_neighbors_log = None
        self.show_tengig_bw_log = None
        self.show_tengig_bw = None     # 1G or None
        self.unknown_mac = []
        self.down_port_with_description = {}    # {port : description}

        self.pagg = "-"
        self.exclude_inf = []  # exclude interface vlans
        self.description_exclude = ["UPLINK", "DOWNLINK", "csg", "pagg", "ACCESS", "MGMT", "MNG", "ME"]

        # database gathered from show_mac show_arp
        self.bs = {}    # mac : {bs_id:, port:, if_vlan: [], vlan: []}
        # {48fd.8e05.6fa7: {bs_id: AL7777, port: Gi0/8, inf: [1000-1005], bs: AL7410}}

        # database gathered from show_description
        self.port_bs = {}  # {port: [bs1, bs2]}       {Gi0/8: [AL7374_7008_7007, ALR746]}
        self.ifvlan_bs = {}  # {int vlan: [bs1, bs2]}    {1000: [AL7374_7008_7007, ALR746]}

        self.lag = {}   # Po1 : [members : Gi0/4 Gi0/5, tag: AU7070 Po1, AU7070 Po1]
        self.removed_info = []  # from def delete_info

        self.commands = []
        self.configuration_log = []

    def show_commands(self):
        self.show_arp_log = self.ssh_conn.send_command(r"show ip arp vrf MA | exclude -|Incomplete")
        self.show_isis_log = self.ssh_conn.send_command(r"show isis hostname | include pagg")
        self.show_isis_neighbors_log = self.ssh_conn.send_command(r"show isis neighbors | include Vl")
        self.show_mac_log = self.ssh_conn.send_command(r"show mac-address-table")
        self.show_description_log = self.ssh_conn.send_command(r"show interfaces description")
        self.show_tengig_bw_log = self.ssh_conn.send_command(r"show interfaces te0/0 | in BW")

    def parse(self, dev, bs_dict, bs_dict_backup):
        csg_mac_log_parse(dev, bs_dict, bs_dict_backup)
        csg_arp_log_parse(dev)
        csg_define_pagg(dev)
        csg_description_parse(dev)
        csg_tengig_bw_parse(dev)

    def delete_info(self, dev):
        csg_delete_info(dev)

    def define_port_bs(self, dev):
        csg_port_bs(dev)

    def lag_member_tag(self, dev):
        csg_lag_member_tag(dev)

    def make_config(self, dev):
        csg_make_config(dev)

    def configure(self, cmd):
        self.configuration_log.append(self.ssh_conn.send_config_set(cmd))

    def commit(self):
        try:
            self.configuration_log.append(self.ssh_conn.save_config())
        except Exception as err_msg:
            self.configuration_log.append(f"COMMIT is OK after msg:{err_msg}")
            self.configuration_log.append(self.ssh_conn.send_command("\n", expect_string=r"#"))

    def reset(self):
        self.connection_status = True
        self.connection_error_msg = None
        self.show_isis_log = None
        self.show_mac_log = None
        self.show_arp_log = None
        self.show_description_log = None
        self.show_isis_neighbors_log = None
        self.show_tengig_bw_log = None
        self.show_tengig_bw = None
        self.pagg = "-"
        self.exclude_inf = []
        self.bs = {}
        self.port_bs = {}
        self.ifvlan_bs = {}
        self.lag = {}
        self.removed_info = []
        self.commands = []
        self.configuration_log = []
        self.down_port_with_description = {}


class PaggXR(CellSiteGateway):

    def __init__(self, ip, host):
        CellSiteGateway.__init__(self, ip, host)
        self.os_type = "cisco_xr"

    def commit(self):
        self.configuration_log.append(self.ssh_conn.commit())
        self.ssh_conn.exit_config_mode()

    def configure(self, cmd):
        self.ssh_conn.send_config_set(cmd)
        self.configuration_log.append(self.ssh_conn.send_command("show configuration"))

    def show_commands(self):
        self.show_arp_log = self.ssh_conn.send_command(r"show arp vrf MA | exclude Interface")
        self.show_description_log = self.ssh_conn.send_command('show interfaces description')
        self.show_mac_log = \
            self.ssh_conn.send_command("show l2vpn forwarding bridge-domain mac-address location 0/0/CPU0")

    def parse(self, dev, bs_dict, bs_dict_backup):
        pagg_arp_log_parse(dev, bs_dict, bs_dict_backup)
        pagg_mac_log_parse(dev)     # после arp
        pagg_description_parse(dev)
        dev.pagg = dev.hostname

    def lag_member_tag(self, dev):
        pagg_lag_member_tag(dev)

    def delete_info(self, dev):
        pagg_delete_info(dev)

    def define_port_bs(self, dev):
        pagg_port_bs(dev)

    def make_config(self, dev):
        pagg_make_config(dev)


class PaggXE(CellSiteGateway):

    def __init__(self, ip, host):
        CellSiteGateway.__init__(self, ip, host)
        self.os_type = "cisco_xe"

    def show_commands(self):
        self.show_arp_log = self.ssh_conn.send_command(r"show ip arp vrf MA | exclude -|Incomplete")
        self.show_mac_log = self.ssh_conn.send_command(r"show mac-address-table dynamic")
        self.show_description_log = self.ssh_conn.send_command(r"show interfaces description")

    def parse(self, dev, bs_dict, bs_dict_backup):
        xe_mac_log_parse(dev, bs_dict, bs_dict_backup)
        xe_description_parse(dev)
        dev.pagg = dev.hostname

    def delete_info(self, dev):
        csg_delete_info(dev)

    def define_port_bs(self, dev):
        csg_port_bs(dev)

    def lag_member_tag(self, dev):
        xe_lag_member_tag(dev)

    def make_config(self, dev):
        xe_make_config(dev)

    def commit(self):
        self.configuration_log.append(self.ssh_conn.save_config())


#######################################################################################
# ------------------------------ def function part -----------------------------------#
#######################################################################################


def get_arguments(arguments):
    settings = {"maxth": 20, "conf": False}
    mt_pattern = re.compile(r"mt([0-9]+)")
    for arg in arguments:
        if "mt" in arg:
            match = re.search(mt_pattern, arg)
            if match and int(match[1]) <= 100:
                settings["maxth"] = int(match[1])
        elif arg == "cfg" or arg == "CFG" or arg == "conf":
            settings["conf"] = True
    
    print("\n"
          f"max threads:...................{settings['maxth']}\n"
          f"config mode:...................{settings['conf']}\n"
          )
    return settings


def get_user_pw():
    with open("psw.yaml") as file:
        user_psw = yaml.load(file, yaml.SafeLoader)

    return user_psw[0], user_psw[1]


def get_device_info(csv_file):
    devs = []
    with open(csv_file, "r") as file:
        for line in file:
            if "#" not in line and len(line) > 5:
                hostname, ip_address, ios = line.strip().split(",")

                if ios in ["ios", "cisco_ios"]:
                    dev = CellSiteGateway(ip=ip_address, host=hostname)
                    devs.append(dev)
                elif ios in ["ios xr", "cisco_xr", "xr"]:
                    dev = PaggXR(ip=ip_address, host=hostname)
                    devs.append(dev)
                elif ios in ["ios xe", "cisco_xe", "xe"]:
                    dev = PaggXE(ip=ip_address, host=hostname)
                    devs.append(dev)
                else:
                    print(f"wrong ios: {ios}")

    return devs


def load_excel(curr_date, curr_time):
    excel_file = input("Enter IP-MAC excel file (by default no excel file is loaded): ")
    result = {}  # mac : bs

    with open("mac_bs_backup.yaml") as f:   
        yaml_file_backup = yaml.load(f, yaml.SafeLoader)

    if excel_file:
        if excel_file.endswith('xlsx'):
            wb = load_workbook(excel_file)
            first_sheet = wb.sheetnames[0]
            sheet = wb[first_sheet]
            x = 2

            while True:
                mac = sheet.cell(row=x, column=4).value
                bs = sheet.cell(row=x, column=1).value
                if bs:
                    x += 1
                    mac_split = mac.split(":")
                    mac_final = "{}{}.{}{}.{}{}".format(mac_split[0], mac_split[1],
                                                        mac_split[2], mac_split[3],
                                                        mac_split[4], mac_split[5])
                    result[mac_final] = bs
                else:
                    break

            with open("mac_bs.yaml", "w") as output_file:
                output_file.write(f"# {curr_date} {curr_time}\n\n")
                for i, j in result.items():
                    output_file.write(f"{i} : {j}\n")

        elif excel_file.endswith('csv'):
            with open(excel_file) as csv_file:
                csv_reader = csv.reader(csv_file, delimiter=',')
                line_count = 0
                for row in csv_reader:
                    if line_count == 0:
                        line_count += 1
                    else:
                        bs = row[0]
                        mac_dd = row[3]
                        mac_split = mac_dd.split(":")
                        mac = "{}{}.{}{}.{}{}".format(mac_split[0], mac_split[1],
                                                      mac_split[2], mac_split[3],
                                                      mac_split[4], mac_split[5])
                        result[mac] = bs

            with open("mac_bs.yaml", "w") as output_file:
                output_file.write(f"# {curr_date} {curr_time}\n\n")
                for i, j in result.items():
                    output_file.write(f"{i} : {j}\n")

        with open("mac_bs_backup.yaml", "a") as file_backup:
            # update backup mac-bs file
            result_keys = result.keys()
            yaml_file_backup_keys = yaml_file_backup.keys()
            diff_keys = [y for y in result_keys if y not in yaml_file_backup_keys]

            file_backup.write(f"\n# {curr_date} {curr_time}\n")
            for diff_key in diff_keys:
                file_backup.write(f"{diff_key} : {result[diff_key]}\n")

        print("mac_bs.yaml, mac_bs_backup.yaml are created")

    else:
        with open("mac_bs.yaml", "r") as file:
            yaml_file = yaml.load(file, yaml.SafeLoader)
            result.update(yaml_file)

        print("previous mac_bs.yaml is used")

    return result, yaml_file_backup


def write_logs(devices, current_time, log_folder, settings):
    failed_conn_count = 0
    unavailable_device = []
    devices_with_cfg = []
    unknown_mac = []
    tag_hostname = {}
    bs_hostname = {}

    export_excel(devices, current_time, log_folder)
    export_excel_split(devices, current_time, log_folder)
    
    conn_msg = log_folder / f"{current_time}_connection_error_msg.txt"
    device_info = log_folder / f"{current_time}_device_info.txt"
    config = log_folder / f"{current_time}_configuration_log.txt"
    commands = log_folder / f"{current_time}_configuration_needed.txt"
    removed = log_folder / f"{current_time}_removed_info.txt"
    tag_hostname_file = log_folder / f"{current_time}_tag_hostname.txt"
    bs_hostname_file = log_folder / f"{current_time}_bs_hosname.txt"
    down_port_with_description = log_folder / f"{current_time}_down_port_with_description.txt"

    conn_msg_file = open(conn_msg, "w")
    device_info_file = open(device_info, "w")
    config_file = open(config, "w")
    commands_file = open(commands, "w")
    removed_file = open(removed, "w")
    down_port_with_description_file = open(down_port_with_description, "w")

    for device in devices:
        if device.connection_status:
            export_device_info(device, device_info_file)  # export device info: show, status, etc
        else:
            failed_conn_count += 1
            conn_msg_file.write("-" * 80 + "\n")
            conn_msg_file.write(f"### {device.hostname} : {device.ip_address} ###\n\n")
            conn_msg_file.write(f"{device.connection_error_msg}\n")
            unavailable_device.append(f"{device.hostname} : {device.ip_address}")
            
        if settings["conf"] and device.commands:
            config_file.write("#" * 80 + "\n")
            config_file.write(f"### {device.hostname} : {device.ip_address} ###\n\n")
            config_file.write("".join(device.configuration_log))
            config_file.write("\n\n")
        elif not settings["conf"] and device.commands:
            commands_file.write(f"### {device.hostname} : {device.ip_address}\n\n")
            commands_file.write("\n".join(device.commands))
            commands_file.write("\n\n\n")
            devices_with_cfg.append(f"{device.hostname},{device.ip_address},{device.os_type}")

        if device.removed_info:
            removed_file.write(f"{device.hostname}\t{' '.join(device.removed_info)}\n")

        if device.unknown_mac:
            unknown_mac.extend(device.unknown_mac)

        if device.down_port_with_description:
            for port, description in device.down_port_with_description.items():
                down_port_with_description_file.write(f"{device.hostname},{device.ip_address},{port},{description}\n")

        # check if a optic bs is connectec via RRL
        for pv in device.port_bs.values():
            if not tag_hostname.get(pv["tag"]):
                tag_hostname[pv["tag"]] = device.hostname
        for mc in device.bs.values():
            bs_hostname[mc["bs_id"]] = device.hostname

    conn_msg_file.close()
    device_info_file.close()
    config_file.close()
    commands_file.close()
    removed_file.close()
    down_port_with_description_file.close()

    if not settings["conf"]:
        config.unlink()
    if all([dev.connection_status is True for dev in devices]):
        conn_msg.unlink()
    if all([not len(dev.down_port_with_description) for dev in devices]):
        down_port_with_description.unlink()
    if not settings["conf"] and devices_with_cfg:
        print("\n" + "-" * 103 + "\n")
        print(f"devices with cfg ({len(devices_with_cfg)}):\n")
        for d in devices_with_cfg:
            print(d)
    if unknown_mac:
        print("\n" + "-" * 103 + "\n")
        print(f"devices with unknown mac ({len(unknown_mac)}):\n")
        for u in unknown_mac:
            print(u)
    if unavailable_device:
        print("\n" + "-" * 103 + "\n")
        print(f"unavailable devices ({len(unavailable_device)}):\n")
        for ud in unavailable_device:
            print(ud)

    # check if a optic bs is connectec via RRL
    print("\n" + "-" * 103 + "\n"
          "bs on multiple devices check:\n")
    with open(tag_hostname_file, "w") as thf:
        for tagk, tagv in tag_hostname.items():
            thf.write(f"{tagk} : {tagv}\n")
    with open(bs_hostname_file, "w") as bhf:
        for bsk, bsv in bs_hostname.items():
            bhf.write(f"{bsk} : {bsv}\n")
    for bsid, bshost in bs_hostname.items():
        for tagid, taghost in tag_hostname.items():
            if bsid in tagid and bshost != taghost:
                print(f"{bsid} is on devices: {bshost}, {taghost}")

    return failed_conn_count


def export_excel(devs, current_time, log_folder):
    filename = log_folder / f"{current_time}_mbh_bs_list_(merged_cells).xlsx"
    wb = Workbook()
    sheet = wb.active
    sheet.append(["PAGG",
                  "CSG hostname",
                  "CSG loopback0",
                  "CSG port",
                  "port tag",
                  "BS",
                  "comments"])
    for dev in devs:
        if dev.connection_status:
            for port, port_info in dev.port_bs.items():
                if dev.lag.get(port):
                    if len(port_info["bs"]) > 0:
                        sheet.append([dev.pagg,
                                      dev.hostname,
                                      dev.ip_address,
                                      f'{port} ({len(dev.lag[port]["members"])} Gps)',
                                      f'{" ".join(set(dev.lag[port]["tag"]))}',
                                      ' '.join(port_info["bs"])])
                    else:
                        sheet.append([dev.pagg,
                                      dev.hostname,
                                      dev.ip_address,
                                      f'{port} ({len(dev.lag[port]["members"])} Gps)',
                                      f'{" ".join(set(dev.lag[port]["tag"]))}',
                                      "",
                                      "no bs"
                                      ])

                else:
                    if any(i in port_info["tag"] for i in ["iot", "IOT", "IoT", "lora", "LORA"]) and \
                            len(port_info["bs"]) == 0:
                        sheet.append([dev.pagg,
                                      dev.hostname,
                                      dev.ip_address,
                                      port,
                                      port_info["tag"],
                                      "-"])

                    elif "Te" in port and dev.show_tengig_bw == "1G":
                        if len(port_info["bs"]) > 0:
                            sheet.append([dev.pagg,
                                          dev.hostname,
                                          dev.ip_address,
                                          f'{port} (1 Gps)',
                                          port_info["tag"],
                                          ' '.join(port_info["bs"])])
                        else:
                            sheet.append([dev.pagg,
                                          dev.hostname,
                                          dev.ip_address,
                                          f'{port} (1 Gps)',
                                          port_info["tag"],
                                          "",
                                          "no bs"])
                    else:
                        if len(port_info["bs"]) > 0:
                            sheet.append([dev.pagg,
                                          dev.hostname,
                                          dev.ip_address,
                                          port,
                                          port_info["tag"],
                                          ' '.join(port_info["bs"])])
                        else:
                            sheet.append([dev.pagg,
                                          dev.hostname,
                                          dev.ip_address,
                                          port,
                                          port_info["tag"],
                                          "",
                                          "no bs"])
        else:
            sheet.append([dev.pagg,
                          dev.hostname,
                          dev.ip_address,
                          "-",
                          "-",
                          "-",
                          "unavailable"])

    # merge same column
    for c in [1, 2, 3]:  # PAGG, CSG hostname, CSG loopback0
        r = 2
        start_r = 0
        while True:
            r += 1
            cell = sheet.cell(row=r, column=c).value
            previous_cell = sheet.cell(row=r - 1, column=c).value
            if cell:
                if cell == previous_cell:
                    start_r += 1
                else:
                    if start_r > 0:
                        sheet.merge_cells(start_row=r - 1 - start_r, start_column=c, end_row=r - 1, end_column=c)
                        sheet.cell(row=r - 1 - start_r, column=c).alignment = styles.Alignment(vertical='center')
                        start_r = 0

            else:
                if start_r > 0:
                    sheet.merge_cells(start_row=r - 1 - start_r, start_column=c, end_row=r - 1, end_column=c)
                    sheet.cell(row=r - 1 - start_r, column=c).alignment = styles.Alignment(vertical='center')
                break

    wb.save(filename)


def export_excel_split(devs, current_time, log_folder):
    filename = log_folder / f"{current_time}_mbh_bs_list.xlsx"
    wb = Workbook()
    sheet = wb.active
    sheet.append(["PAGG",
                  "CSG hostname",
                  "CSG loopback0",
                  "CSG port",
                  "port tag",
                  "BS",
                  "comments"])
    for dev in devs:
        if dev.connection_status:
            for port, port_info in dev.port_bs.items():
                if dev.lag.get(port):
                    if len(port_info["bs"]) > 0:
                        sheet.append([dev.pagg,
                                      dev.hostname,
                                      dev.ip_address,
                                      f'{port} ({len(dev.lag[port]["members"])} Gps)',
                                      f'{" ".join(set(dev.lag[port]["tag"]))}',
                                      ' '.join(port_info["bs"])])
                    else:
                        sheet.append([dev.pagg,
                                      dev.hostname,
                                      dev.ip_address,
                                      f'{port} ({len(dev.lag[port]["members"])} Gps)',
                                      f'{" ".join(set(dev.lag[port]["tag"]))}',
                                      "",
                                      "no bs"
                                      ])

                else:
                    if any(i in port_info["tag"] for i in ["iot", "IOT", "IoT", "lora", "LORA"]) and \
                            len(port_info["bs"]) == 0:
                        sheet.append([dev.pagg,
                                      dev.hostname,
                                      dev.ip_address,
                                      port,
                                      port_info["tag"],
                                      "-"])

                    elif "Te" in port and dev.show_tengig_bw == "1G":
                        if len(port_info["bs"]) > 0:
                            sheet.append([dev.pagg,
                                          dev.hostname,
                                          dev.ip_address,
                                          f'{port} (1 Gps)',
                                          port_info["tag"],
                                          ' '.join(port_info["bs"])])
                        else:
                            sheet.append([dev.pagg,
                                          dev.hostname,
                                          dev.ip_address,
                                          f'{port} (1 Gps)',
                                          port_info["tag"],
                                          "",
                                          "no bs"])
                    else:
                        if len(port_info["bs"]) > 0:
                            sheet.append([dev.pagg,
                                          dev.hostname,
                                          dev.ip_address,
                                          port,
                                          port_info["tag"],
                                          ' '.join(port_info["bs"])])
                        else:
                            sheet.append([dev.pagg,
                                          dev.hostname,
                                          dev.ip_address,
                                          port,
                                          port_info["tag"],
                                          "",
                                          "no bs"])
        else:
            sheet.append([dev.pagg,
                          dev.hostname,
                          dev.ip_address,
                          "-",
                          "-",
                          "-",
                          "unavailable"])

    wb.save(filename)


def export_device_info(dev, export_file):
    export_file.write("#" * 80 + "\n")
    export_file.write(f"### {dev.hostname} : {dev.ip_address} ###\n\n")

    if dev.show_isis_log:
        export_file.write("-" * 80 + "\n")
        export_file.write("device.show_isis_log\n\n")
        export_file.write(dev.show_isis_log)
        export_file.write("\n\n")

    if dev.show_mac_log:
        export_file.write("-" * 80 + "\n")
        export_file.write("device.show_mac_log\n\n")
        export_file.write(dev.show_mac_log)
        export_file.write("\n\n")

    if dev.show_arp_log:
        export_file.write("-" * 80 + "\n")
        export_file.write("device.show_arp_log\n\n")
        export_file.write(dev.show_arp_log)
        export_file.write("\n\n")

    if dev.show_description_log:
        export_file.write("-" * 80 + "\n")
        export_file.write("device.show_description_log\n\n")
        export_file.write(dev.show_description_log)
        export_file.write("\n\n")

    if dev.show_isis_neighbors_log:
        export_file.write("-" * 80 + "\n")
        export_file.write("device.show_isis_neighbors_log\n\n")
        export_file.write(dev.show_isis_neighbors_log)
        export_file.write("\n\n")

    export_file.write("-" * 80 + "\n")
    export_file.write("device.bs\n\n")
    export_file.write(pformat(dev.bs))
    export_file.write("\n\n")

    export_file.write("-" * 80 + "\n")
    export_file.write("device.port_bs\n\n")
    export_file.write(pformat(dev.port_bs))
    export_file.write("\n\n")

    export_file.write("-" * 80 + "\n")
    export_file.write("device.ifvlan_bs\n\n")
    export_file.write(pformat(dev.ifvlan_bs))
    export_file.write("\n\n")

    export_file.write("-" * 80 + "\n")
    export_file.write("device.lag\n\n")
    export_file.write(pformat(dev.lag))
    export_file.write("\n\n")

    export_file.write("-" * 80 + "\n")
    export_file.write("device.exclude_inf\n\n")
    export_file.write(" ".join(dev.exclude_inf))
    export_file.write("\n\n")

    export_file.write("-" * 80 + "\n")
    export_file.write("device.commands\n\n")
    export_file.write(pformat(dev.commands))
    export_file.write("\n\n")

    export_file.write("-" * 80 + "\n")
    export_file.write("device.down_port_with_description\n\n")
    export_file.write(pformat(dev.down_port_with_description))
    export_file.write("\n\n")


#######################################################################################
# ------------------------------ get bs port -----------------------------------------#
#######################################################################################


def define_inf_exclude(dev):
    dev.exclude_inf.extend([str(i) for i in range(1080, 1199)])  # MA BGP
    dev.exclude_inf.extend([str(i) for i in range(4000, 4099)])  # MW MGMT
    dev.exclude_inf.extend([str(i) for i in range(2020, 2099)])  # SMART METERING

    if dev.os_type == "cisco_ios":
        for line in dev.show_isis_neighbors_log.splitlines():
            match = re.search(r".*L2 +Vl(\d+) +", line)     # akta-040001-pag L2 Vl(200) 10.238.121.65
            if match:
                dev.exclude_inf.append(match[1])


def csg_mac_log_parse(dev, bs_dict, bs_dict_backup):
    pattern = re.compile(r"(\d+)\s+(\w{4}\.\w{4}\.\w{4})\s+DYNAMIC\s+(\S+)")
    # (3001)    (48fd.8e05.6fa7)    DYNAMIC     (Gi0/8)
    for line in dev.show_mac_log.splitlines():
        match = re.search(pattern, line)
        if match:
            vlan = match[1]     # 3001
            mac = match[2]      # 48fd.8e05.6fa7
            port = match[3]     # Gi0/8
            if bs_dict.get(mac):
                bs = bs_dict[mac]
            else:
                if bs_dict_backup.get(mac):
                    bs = bs_dict_backup[mac]
                else:
                    bs = mac

            if dev.bs.get(mac):
                dev.bs[mac]["vlan"].append(vlan)
            else:
                dev.bs[mac] = {"bs_id": bs,
                               "port": port,
                               "if_vlan": [],
                               "vlan": [vlan]}
                               
            if "Po" in port:
                dev.lag[port] = {"members": [], "tag": []}


def xe_mac_log_parse(dev, bs_dict, bs_dict_backup):
    pattern = re.compile(r"(\d+) +(\w{4}\.\w{4}\.\w{4}) +DYNAMIC +(Gi\d/\d/\d)")
    # (1000)  (18de.d7aa.7264)  DYNAMIC  (Gi0/5/6).Efp1000
    for line in dev.show_mac_log.splitlines():
        match = re.search(pattern, line)
        if match:
            vlan = match[1]     # продублировать в if_vlan
            mac = match[2]
            port = match[3]
            
            if bs_dict.get(mac):
                bs = bs_dict[mac]
            else:
                if bs_dict_backup.get(mac):
                    bs = bs_dict_backup[mac]
                else:
                    bs = mac
                
            if dev.bs.get(mac):
                dev.bs[mac]["vlan"].append(vlan)
                dev.bs[mac]["if_vlan"].append(vlan)
            else:
                dev.bs[mac] = {"bs_id": bs,
                               "port": port,
                               "if_vlan": [vlan],
                               "vlan": [vlan]}
                               
            if "Po" in port:
                dev.lag[port] = {"members": [], "tag": []}


def pagg_mac_log_parse(dev):
    # only for bvi
    pattern = re.compile(r"(\w{4}\.\w{4}\.\w{4}) +dynamic +(Gi[0-9/]+)\.(\d+)")
    # (688f.845f.bf7d) dynamic (Gi0/0/0/18).(1000)   N/A  17 Mar 21:05:41  N/A

    if dev.show_mac_log:
        for line in dev.show_mac_log.splitlines():
            match = re.search(pattern, line)
            if match:
                mac = match[1]
                port = match[2]
                bvi = match[3]      # для проверки

                if dev.bs.get(mac):
                    dev.bs[mac]["port"] = port

                    if bvi not in dev.bs[mac]["if_vlan"]:
                        print(f"{dev.hostname:39}pagg_mac_log_parse: {bvi} is not in dev.bs.ifvlan")

                else:
                    print(f"{dev.hostname:39}pagg_mac_log_parse: {mac} is not in dev.bs")


def csg_arp_log_parse(dev):
    pattern = re.compile(r"Internet\s+\d+\.\d+\.\d+\.\d+\s+\d+\s+(\w{4}\.\w{4}\.\w{4})\s+ARPA\s+Vlan(\d+)")
    # Internet  10.165.161.87          11   (d849.0b95.af44)  ARPA   Vlan(1000)
    for line in dev.show_arp_log.splitlines():
        match = re.search(pattern, line)  # ip mac inf_vlan
        if match:
            mac = match[1]  # d849.0b95.af44
            inf = match[2]  # 1000 (without Vlan)
            if dev.bs.get(mac):
                dev.bs[mac]["if_vlan"].append(inf)
            else:
                print(f"{dev.hostname:39}arp_log_parse - {mac} not in MAC table")


def pagg_arp_log_parse(dev, bs_dict, bs_dict_backup):
    pattern = re.compile(r"(\w{4}\.\w{4}\.\w{4}) +Dynamic +ARPA +([-A-Za-z]+)([0-9/]+)\.(\d+)$")
    # 10.146.56.1     00:02:06   (883f.d304.e2a1)  Dynamic    ARPA  (GigabitEthernet)(0/0/0/5).(1080)
    # 10.164.24.243   00:02:11   (845b.1260.9241)  Dynamic    ARPA  (Bundle-Ether)(10).(1004)

    pattern_bvi = re.compile(r"(\w{4}\.\w{4}\.\w{4}) +Dynamic +ARPA +BVI(\d+)")
    # 10.165.192.178  00:00:48   (d849.0b8a.dcd1)  Dynamic    ARPA  BVI(1000)

    for line in dev.show_arp_log.splitlines():
        match = re.search(pattern, line)
        match_bvi = re.search(pattern_bvi, line)
        if match:
            mac = match[1]              # 883f.d304.e2a1
            port_ethernet = match[2]    # GigabitEthernet
            port_number = match[3]      # 0/0/0/5
            vlan = match[4]             # 1080
            
            if bs_dict.get(mac):
                bs = bs_dict[mac]
            else:
                if bs_dict_backup.get(mac):
                    bs = bs_dict_backup[mac]
                else:
                    bs = mac

            if port_ethernet == "Bundle-Ether":
                port_ethernet = "BE"
            elif port_ethernet == "TenGigE":
                port_ethernet = "Te"
            elif port_ethernet == "GigabitEthernet":
                port_ethernet = "Gi"
            else:
                print(f"{dev.hostname:39}pagg_arp_log_parse: Gi,Te,Be not in {port_ethernet}")

            if dev.bs.get(mac):
                dev.bs[mac]["vlan"].append(vlan)
            else:
                dev.bs[mac] = {"port": f'{port_ethernet}{port_number}',
                               "if_vlan": [],
                               "vlan": [vlan],
                               "bs_id": bs}
                               
            if port_ethernet == "BE":
                dev.lag[f"{port_ethernet}{port_number}"] = {"members": [], "tag": []}

        if match_bvi:
            mac = match_bvi[1]
            bvi = match_bvi[2]

            if bs_dict.get(mac):
                bs = bs_dict[mac]
            else:
                if bs_dict_backup.get(mac):
                    bs = bs_dict_backup[mac]
                else:
                    bs = mac

            if dev.bs.get(mac):
                dev.bs[mac]["if_vlan"].append(bvi)
            else:
                dev.bs[mac] = {"port": '',
                               "if_vlan": [bvi],
                               "vlan": [],
                               "bs_id": bs}


def csg_define_pagg(dev):
    pattern = re.compile(r"[0-9.]{14} ([a-z.]+-\d+-pagg-\d)")
    for line in dev.show_isis_log.splitlines():
        match = re.search(pattern, line)
        if match:
            dev.pagg = match[1]


def csg_description_parse(dev):
    pattern_port = re.compile(r"((?:Gi|Te|Po)\S+)\s+up\s+up\s*(.*)")    # (Gi0/6) up up (AK7137 BS: ALG005 AK7160)
    pattern_port_tag_bs = re.compile(r"(?:(.*)\s)?BS:\s?(.*)")          # (AK7137) BS: (ALG005 AK7160)
    pattern_inf = re.compile(r"Vl(\d+)\s+up\s+up\s*(.*)")               # Vl(1000) up up (ABIS BS: ALG005 AK7160)
    pattern_inf_tag_bs = re.compile(r"(?:.*\s)?BS:\s?(.*)")             # ABIS BS: (ALG005 AK7160)

    for line in dev.show_description_log.splitlines():
        match_port = re.search(pattern_port, line)
        match_inf = re.search(pattern_inf, line)
        if match_port:
            port = match_port[1]
            description = match_port[2]
            if not any(i in line for i in dev.description_exclude):
                if "BS:" in description:
                    match_port_tag_bs = re.search(pattern_port_tag_bs, description)
                    if match_port_tag_bs:
                        tag = match_port_tag_bs[1]
                        bs = match_port_tag_bs[2]
                        dev.port_bs[port] = {"tag": f'{tag if tag else ""}',
                                             "bs": [],
                                             "new_bs_description": "",
                                             "current_bs_description": bs,
                                             "bs_on_description": []}
                    else:
                        print(f"{dev.hostname:39}{match_port_tag_bs} re match error")
                else:
                    if len(description) > 0:
                        dev.port_bs[port] = {"tag": description,
                                             "bs": [],
                                             "new_bs_description": "",
                                             "current_bs_description": "",
                                             "bs_on_description": []}
                    else:
                        dev.port_bs[port] = {"tag": "",
                                             "bs": [],
                                             "new_bs_description": "",
                                             "current_bs_description": "",
                                             "bs_on_description": []}
        elif match_inf:
            inf = match_inf[1]
            description = match_inf[2]
            tag = None
            if inf not in dev.exclude_inf and not any(i in line for i in dev.description_exclude):
                for m in ["ABIS", "IUB", "OAM", "S1U", "S1MME", "X2", "S1C"]:
                    if m in description:
                        tag = m
                        break
                if tag is None:
                    print(f"{dev.hostname:39}no ABIS,X2,IUB,S1MME,S1U,S1C,OAM in description interface vlan{inf}")
                if "BS:" in description:
                    match_inf_tag_bs = re.search(pattern_inf_tag_bs, description)
                    if match_inf_tag_bs:
                        bs = match_inf_tag_bs[1]
                        dev.ifvlan_bs[inf] = {"tag": tag,
                                              "bs": [],
                                              "new_bs_description": "",
                                              "current_bs_description": bs,
                                              "bs_on_description": []}
                    else:
                        print(f"{dev.hostname:39}{match_inf_tag_bs} re match error")
                else:
                    if len(description) > 0 and tag is not None:
                        dev.ifvlan_bs[inf] = {"tag": tag,
                                              "bs": [],
                                              "new_bs_description": "",
                                              "current_bs_description": "",
                                              "bs_on_description": []}
                    else:
                        print(f"{dev.hostname:39}no description interface vlan{inf}")


def xe_description_parse(dev):
    pattern_port = re.compile(r"((?:Gi|Po)\S+)\s+up\s+up\s*(.*)")
    pattern_port_tag_bs = re.compile(r"(?:(.*)\s)?BS:\s?(.*)")
    pattern_inf = re.compile(r"BD(\d+)\s+up\s+up\s*(.*)")       # BD(1000)
    pattern_inf_tag_bs = re.compile(r"(?:.*\s)?BS:\s?(.*)")     # ABIS BS: (ALG005 AK7160)
    
    for line in dev.show_description_log.splitlines():
        match_port = re.search(pattern_port, line)
        match_inf = re.search(pattern_inf, line)
        if match_port:
            port = match_port[1]
            description = match_port[2]
            if not any(i in line for i in dev.description_exclude):
                if "BS:" in description:
                    match_port_tag_bs = re.search(pattern_port_tag_bs, description)
                    if match_port_tag_bs:
                        tag = match_port_tag_bs[1]
                        bs = match_port_tag_bs[2]
                        dev.port_bs[port] = {"tag": f'{tag if tag else ""}',
                                             "bs": [],
                                             "new_bs_description": "",
                                             "current_bs_description": bs,
                                             "bs_on_description": []}
                    else:
                        print(f"{dev.hostname:39}{match_port_tag_bs} re match error")
                else:
                    if len(description) > 0:
                        dev.port_bs[port] = {"tag": description,
                                             "bs": [],
                                             "new_bs_description": "",
                                             "current_bs_description": "",
                                             "bs_on_description": []}
                    else:
                        dev.port_bs[port] = {"tag": "",
                                             "bs": [],
                                             "new_bs_description": "",
                                             "current_bs_description": "",
                                             "bs_on_description": []}
        elif match_inf:
            inf = match_inf[1]
            description = match_inf[2]
            tag = None
            if inf not in dev.exclude_inf and not any(i in line for i in dev.description_exclude):
                for m in ["ABIS", "IUB", "OAM", "S1U", "S1MME", "X2", "S1C"]:
                    if m in description:
                        tag = m
                        break
                if tag is None:
                    print(f"{dev.hostname:39}no ABIS,X2,IUB,S1MME,S1U,S1C,OAM in description interface vlan{inf}")
                if "BS:" in description:
                    match_inf_tag_bs = re.search(pattern_inf_tag_bs, description)
                    if match_inf_tag_bs:
                        bs = match_inf_tag_bs[1]
                        dev.ifvlan_bs[inf] = {"tag": tag,
                                              "bs": [],
                                              "new_bs_description": "",
                                              "current_bs_description": bs,
                                              "bs_on_description": []}
                    else:
                        print(f"{dev.hostname:39}{match_inf_tag_bs} re match error")
                else:
                    if len(description) > 0 and tag is not None:
                        dev.ifvlan_bs[inf] = {"tag": tag,
                                              "bs": [],
                                              "new_bs_description": "",
                                              "current_bs_description": "",
                                              "bs_on_description": []}
                    else:
                        print(f"{dev.hostname:39}no description interface vlan{inf}")


def pagg_description_parse(dev):
    pattern = re.compile(r"((?:Gi|Te|BE)[0-9/]+)\s+up\s+up\s*(.*)$")
    # (Gi0/0/0/5)     up  up  (AU7104 BS: ZHA012)
    # Gi0/0/0/5.1000  up  up  AU7104
    pattern_tag_bs = re.compile(r"(?:(.*)\s)?BS:\s?(.*)")
    # (AU7104) BS: (ZHA012)
    pattern_bvi = re.compile(r"BV(\d+) +up +up *(.*)$")
    pattern_bvi_tag_bs = re.compile(r"(?:.*\s)?BS:\s?(.*)")

    for line in dev.show_description_log.splitlines():
        match = re.search(pattern, line)
        match_bvi = re.search(pattern_bvi, line)
        if match:
            port = match[1]
            description = match[2]
            if not any(i in line for i in dev.description_exclude):
                if "BS:" in description:
                    match_port_tag_bs = re.search(pattern_tag_bs, description)
                    if match_port_tag_bs:
                        tag = match_port_tag_bs[1]
                        bs = match_port_tag_bs[2]
                        dev.port_bs[port] = {"tag": f'{tag if tag else ""}',
                                             "bs": [],
                                             "new_bs_description": "",
                                             "current_bs_description": bs,
                                             "bs_on_description": []}
                    else:
                        print(f"{dev.hostname:39}{match_port_tag_bs} re match error")
                else:
                    if len(description) > 0:
                        dev.port_bs[port] = {"tag": description,
                                             "bs": [],
                                             "new_bs_description": "",
                                             "current_bs_description": "",
                                             "bs_on_description": []}
                    else:
                        dev.port_bs[port] = {"tag": "",
                                             "bs": [],
                                             "new_bs_description": "",
                                             "current_bs_description": "",
                                             "bs_on_description": []}

        elif match_bvi:
            bvi = match_bvi[1]
            description = match_bvi[2]
            tag = None
            if bvi not in dev.exclude_inf and not any(i in line for i in dev.description_exclude):
                for m in ["ABIS", "IUB", "OAM", "S1U", "S1MME", "X2", "S1C"]:
                    if m in description:
                        tag = m
                        break
                if tag is None:
                    print(f"{dev.hostname:39}no ABIS,X2,IUB,S1MME,S1U,S1C,OAM in description interface vlan{bvi}")
                if "BS:" in description:
                    match_bvi_tag_bs = re.search(pattern_bvi_tag_bs, description)
                    if match_bvi_tag_bs:
                        bs = match_bvi_tag_bs[1]
                        dev.ifvlan_bs[bvi] = {"tag": tag,
                                              "bs": [],
                                              "new_bs_description": "",
                                              "current_bs_description": bs,
                                              "bs_on_description": []}
                    else:
                        print(f"{dev.hostname:39}{match_bvi_tag_bs} re match error")
                else:
                    if len(description) > 0 and tag is not None:
                        dev.ifvlan_bs[bvi] = {"tag": tag,
                                              "bs": [],
                                              "new_bs_description": "",
                                              "current_bs_description": "",
                                              "bs_on_description": []}
                    else:
                        print(f"{dev.hostname:39}no description interface vlan{bvi}")


def csg_tengig_bw_parse(dev):
    if "BW 1000000 Kbit" in dev.show_tengig_bw_log:
        dev.show_tengig_bw = "1G"


def csg_lag_member_tag(dev):
    if dev.lag:
        pattern = re.compile(r"((?:Gi|Te)\S+)\s+up\s+up\s*(.*)")
        for port in dev.lag:
            log = dev.ssh_conn.send_command(f"show etherchannel {port[2:]} summary | include LACP")     # номер порта
            lag_members = re.findall(r"(?:Gi|Te)\d/\d{1,2}", log)
            dev.lag[port]["members"].extend(lag_members)
            for line in dev.show_description_log.splitlines():
                match = re.search(pattern, line)
                if match:
                    port_gi = match[1]
                    tag = match[2]
                    if port_gi in lag_members:
                        dev.lag[port]["tag"].append(tag)
    
    
def xe_lag_member_tag(dev):
    if dev.lag:
        pattern = re.compile(r"((?:Gi)\S+)\s+up\s+up\s*(.*)")
        for port in dev.lag:
            log = dev.ssh_conn.send_command(f"show etherchannel summary | include {port}")
            lag_members = re.findall(r"(?:Gi)\d/\d/\d", log)
            dev.lag[port]["members"].extend(lag_members)
            for line in dev.show_description_log.splitlines():
                match = re.search(pattern, line)
                if match:
                    port_gi = match[1]
                    tag = match[2]
                    if port_gi in lag_members:
                        dev.lag[port]["tag"].append(tag)
            

def pagg_lag_member_tag(dev):
    if dev.lag:
        pattern = re.compile(r"(Gi[0-9/]+)\s+up\s+up\s*(.*)$")
        for port in dev.lag:
            log = dev.ssh_conn.send_command(f"show bundle {port} | include Local") 
            lag_members = re.findall(r"(?:Gi|Te)\d/\d/\d/\d{1,2}", log)
            dev.lag[port]["members"].extend(lag_members)
            for line in dev.show_description_log.splitlines():
                match = re.search(pattern, line)
                if match:
                    port_gi = match[1]
                    tag = match[2]
                    if port_gi in lag_members:
                        dev.lag[port]["tag"].append(tag)
                        

def csg_delete_info(dev):
    delete_mac = []
    dev_exclude = ["shor.asta-032001-csg-1", "esil.koks-025001-csg-1"]  # устройство на которых БС прописана как L2 
    
    for mac, bs_info in dev.bs.items():
        if len(bs_info["bs_id"]) == 14:     # удалить все неопределенные MAC, 0046.4bb4.8f76=14
            delete_mac.append(mac)
            if mac in dev.show_arp_log and len(bs_info["vlan"]) > 2:
                print(f"{dev.hostname:39}Unknown MAC: {mac}")
                dev.unknown_mac.append(mac)

        elif mac not in dev.show_arp_log and dev.hostname not in dev_exclude:
            delete_mac.append(mac)

    if delete_mac:
        for i in set(delete_mac):
            dev.removed_info.extend(dev.bs[i]['vlan'])
            del dev.bs[i]
            
    if dev.lag:
        for lag_info in dev.lag.values():
            for port in lag_info["members"]:
                if dev.port_bs.get(port):
                    del dev.port_bs[port]


def pagg_delete_info(dev):
    delete_mac = []
    
    for mac, bs_info in dev.bs.items():
        if len(bs_info["bs_id"]) == 14:     # удалить все неопределенные MAC, 0046.4bb4.8f76=14
            delete_mac.append(mac)
            if len(bs_info["vlan"]) > 2:
                print(f"{dev.hostname:39}Unknown MAC: {mac}")
                dev.unknown_mac.append(mac)

    if delete_mac:
        for i in set(delete_mac):
            dev.removed_info.extend(dev.bs[i]['vlan'])
            del dev.bs[i]
            
    if dev.lag:
        for lag_info in dev.lag.values():
            for port in lag_info["members"]:
                if dev.port_bs.get(port):
                    del dev.port_bs[port]


def description_bs_parse(dev):
    for port, port_info in dev.port_bs.items():
        if port_info["current_bs_description"]:
            for i in port_info["current_bs_description"].split():
                if "_" in i:  # AU7311_7069
                    city = i[:2]  # AU
                    for j in i.split("_"):  # [AU7311, 7069]
                        if city in j:
                            dev.port_bs[port]["bs_on_description"].append(j)
                        else:
                            dev.port_bs[port]["bs_on_description"].append(f"{city}{j}")
                else:
                    dev.port_bs[port]["bs_on_description"].append(i)  # ALR100

    for inf, inf_info in dev.ifvlan_bs.items():
        if inf_info["current_bs_description"]:
            for i in inf_info["current_bs_description"].split():
                if "_" in i:
                    city = i[:2]
                    for j in i.split("_"):
                        if city in j:
                            dev.ifvlan_bs[inf]["bs_on_description"].append(j)
                        else:
                            dev.ifvlan_bs[inf]["bs_on_description"].append(f"{city}{j}")
                else:
                    dev.ifvlan_bs[inf]["bs_on_description"].append(i)


def csg_port_bs(dev):
    for bs_info in dev.bs.values():
        port = bs_info["port"]
        bs = bs_info["bs_id"]
        ifvlanlist = bs_info["if_vlan"]

        if dev.port_bs.get(port):
            dev.port_bs[port]["bs"].append(bs)
        else:
            if dev.hostname not in ["asta-032001-csg-2", "atba.koks-025001-csg-1"]:
                print(f"{dev.hostname:39}{port} not in port_bs dict")

        for ifvlan in ifvlanlist:
            if dev.ifvlan_bs.get(ifvlan):
                dev.ifvlan_bs[ifvlan]["bs"].append(bs)
            else:
                if dev.hostname not in ["rudn-005006-csg-1", "zato.kost-055002-csg-1", "oktb.ustk-001034-csg-1"]:
                    print(f"{dev.hostname:39}{ifvlan} vlan not in ifvlan_bs dict")


def pagg_port_bs(dev):
    for bs_info in dev.bs.values():
        port = bs_info["port"]
        bs = bs_info["bs_id"]
        bvi_list = bs_info["if_vlan"]

        if dev.port_bs.get(port):
            dev.port_bs[port]["bs"].append(bs)
        else:
            print(f"{dev.hostname:39}port not in port_bs dict")

        for bvi in bvi_list:
            if dev.ifvlan_bs.get(bvi):
                dev.ifvlan_bs[bvi]["bs"].append(bs)
            else:
                print(f"{dev.hostname:39}{bvi} vlan not in ifvlan_bs dict")


def shorten_bs(dev):
    pattern = re.compile(r"^([A-Z]{2})(\d{4})")  # (AL)(7374)
    for port, port_info in dev.port_bs.items():
        city_bs = {"others": []}  # AL:[7341,7000] AS:[7007,7000] others:[ALR734,TEST_BS]
        bs_desc = []  # [AL7374_7008_7007, AS7374_7375, ALR746]

        for bs in port_info["bs"]:
            match = re.search(pattern, bs)
            if match:
                region = match[1]  # AL
                bs_number = match[2]  # 7341
                if city_bs.get(region):
                    city_bs[region].append(bs_number)  # AL: [7374, 7000], AS: [7007, 7008]
                else:
                    city_bs[region] = [bs_number]  # AL: [7374], AS: [7007]
            else:
                city_bs["others"].append(bs)  # "others": [ALR734, AL100, TEST_BS]

        for i, j in city_bs.items():
            if i != "others":
                bs_desc.append(f"{i}{'_'.join(j)}")  # ["AL7374_7008_7007"]
            elif i == "others" and len(j) > 0:
                bs_desc.extend(j)  # ["AL7374_7008_7007", "ALR734", "AL100", "TEST_BS"]

        new_bs_description = " ".join(bs_desc)
        dev.port_bs[port]["new_bs_description"] = new_bs_description

        if len(new_bs_description) > 200:
            print(f"{dev.hostname:39}{port}: description is longer than 200")

    for inf, inf_info in dev.ifvlan_bs.items():
        city_bs = {"others": []}
        bs_desc = []
        for bs in inf_info["bs"]:
            match = re.search(pattern, bs)
            if match:
                region = match[1]
                bs_number = match[2]
                if city_bs.get(region):
                    city_bs[region].append(bs_number)
                else:
                    city_bs[region] = [bs_number]
            else:
                city_bs["others"].append(bs)
        for i, j in city_bs.items():
            if i != "others":
                bs_desc.append(f"{i}{'_'.join(j)}")
            elif i == "others" and len(j) > 0:
                bs_desc.extend(j)

        dev.ifvlan_bs[inf]["new_bs_description"] = " ".join(bs_desc)


def csg_make_config(dev):
    for port, port_info in dev.port_bs.items():
        if set(port_info["bs"]) != set(port_info["bs_on_description"]):
            dev.commands.append(f"interface {port}")
            dev.commands.append(f"description {port_info['tag']} BS: {port_info['new_bs_description']}")

    for inf, inf_info in dev.ifvlan_bs.items():
        if set(inf_info["bs"]) != set(inf_info["bs_on_description"]):
            dev.commands.append(f"interface Vlan{inf}")
            dev.commands.append(f"description {inf_info['tag']} BS: {inf_info['new_bs_description']}")


def xe_make_config(dev):
    for port, port_info in dev.port_bs.items():
        if set(port_info["bs"]) != set(port_info["bs_on_description"]):
            dev.commands.append(f"interface {port}")
            dev.commands.append(f"description {port_info['tag']} BS: {port_info['new_bs_description']}")

    for inf, inf_info in dev.ifvlan_bs.items():
        if set(inf_info["bs"]) != set(inf_info["bs_on_description"]):
            dev.commands.append(f"interface BDI{inf}")
            dev.commands.append(f"description {inf_info['tag']} BS: {inf_info['new_bs_description']}")


def pagg_make_config(dev):
    for port, port_info in dev.port_bs.items():
        if set(port_info["bs"]) != set(port_info["bs_on_description"]):
            dev.commands.append(f"interface {port} description {port_info['tag']} BS: {port_info['new_bs_description']}"
                                )

    for bvi, bvi_info in dev.ifvlan_bs.items():
        if set(bvi_info["bs"]) != set(bvi_info["bs_on_description"]):
            dev.commands.append(f"interface BVI{bvi}")
            dev.commands.append(f"description {bvi_info['tag']} BS: {bvi_info['new_bs_description']}")


def configure(dev, settings):
    if settings["conf"]:
        if len(dev.commands) > 0:
            dev.configure(dev.commands)
            dev.commit()
        else:
            print(f"{dev.hostname:39}cfg is not needed")
    else:
        if len(dev.commands) > 0:
            print(f"{dev.hostname:39}cfg is needed")


def down_port_with_description(dev):
    # find ports in down state with description 
    pattern = re.compile(r"(\S+)\s+(?:down|admin down|admin-down)\s+(?:down|admin down|admin-down)\s+(\S+.*)")
    for line in dev.show_description_log.splitlines():
        match = re.search(pattern, line)
        if match:
            port = match[1]
            description = match[2]
            if "Vl" not in port and "Don `t open" not in description and "Don't open" not in description and description != "UPLINK":    # exclude Vlan Interface 
                dev.down_port_with_description[port] = description


#######################################################################################
# ------------------------------              ----------------------------------------#
#######################################################################################

def connect_device(my_username, my_password, dev_queue, bs_dict, bs_dict_backup, settings):
    while True:
        dev = dev_queue.get()
        i = 0
        while True:
            try:
                # print(f"{device.hostname:23}{device.ip_address:16}")
                dev.ssh_conn = ConnectHandler(device_type=dev.os_type, ip=dev.ip_address,
                                              username=my_username, password=my_password)
                dev.show_commands()
                define_inf_exclude(dev)
                dev.parse(dev, bs_dict, bs_dict_backup)
                down_port_with_description(dev)
                dev.lag_member_tag(dev)
                dev.delete_info(dev)
                description_bs_parse(dev)
                dev.define_port_bs(dev)
                shorten_bs(dev)
                dev.make_config(dev)
                configure(dev, settings)
                dev.ssh_conn.disconnect()
                dev_queue.task_done()
                break

            except NetMikoTimeoutException as err_msg:
                dev.connection_status = False
                dev.connection_error_msg = str(err_msg)
                print(f"{dev.hostname:23}{dev.ip_address:16}timeout")
                dev_queue.task_done()
                break
                 
            except Exception as err_msg:
                if i == 2:  # tries
                    dev.connection_status = False
                    dev.connection_error_msg = str(err_msg)
                    print(f"{dev.hostname:23}{dev.ip_address:16}{'BREAK connection failed':20} i={i}")
                    dev_queue.task_done()
                    break
                else:
                    i += 1
                    dev.reset()
                    # print(f"{dev.hostname:23}{dev.ip_address:16}ERROR connection failed i={i}")
                    time.sleep(5)
'''
def connect_device(my_username, my_password, dev_queue, bs_dict, bs_dict_backup, settings):
    dev = dev_queue.get()
    dev.ssh_conn = ConnectHandler(device_type=dev.os_type, ip=dev.ip_address,
                                  username=my_username, password=my_password)
    dev.show_commands()
    define_inf_exclude(dev)
    dev.parse(dev, bs_dict, bs_dict_backup)
    dev.lag_member_tag(dev)
    dev.delete_info(dev)
    description_bs_parse(dev)
    dev.define_port_bs(dev)
    shorten_bs(dev)
    dev.make_config(dev)
    configure(dev, settings)
    dev.ssh_conn.disconnect()
    dev_queue.task_done()
'''


#######################################################################################
# ------------------------------ test        -----------------------------------------#
#######################################################################################

def test_connect_dev(dev, settings):
    if settings["os_type"] == "cisco_ios":
        with open("test_arp.txt", "r") as arp:
            dev.show_arp_log = arp.read()
        with open("test_descrip.txt", "r") as descr:
            dev.show_description_log = descr.read()
        with open("test_isis_pagg.txt", "r") as isis_host:
            dev.show_isis_log = isis_host.read()
        with open("test_isis_neig.txt", "r") as isis_neigh:
            dev.show_isis_neighbors_log = isis_neigh.read()
        with open("test_mac.txt", "r") as mac:
            dev.show_mac_log = mac.read()
        with open("test_tengig.txt", "r") as ten:
            dev.show_tengig_bw_log = ten.read()

    elif settings["os_type"] == "cisco_xr":
        with open("test_pagg_arp.txt", "r") as arp:
            dev.show_arp_log = arp.read()
        with open("test_pagg_description.txt", "r") as descr:
            dev.show_description_log = descr.read()


def test_connect(dev_queue, settings, bs_dict, bs_dict_backup):
    dev = dev_queue.get()
    test_connect_dev(dev, settings)
    define_inf_exclude(dev)
    dev.parse(dev, bs_dict, bs_dict_backup)
    dev.lag_member_tag(dev)
    dev.delete_info(dev)
    description_bs_parse(dev)
    dev.define_port_bs(dev)
    shorten_bs(dev)
    dev.make_config(dev)
    configure(dev, settings)
    dev_queue.task_done()


def test_connect2(my_username, my_password, dev_queue, bs_dict, settings):
    dev = dev_queue.get()
    dev.ssh_conn = ConnectHandler(device_type=dev.os_type, ip=dev.ip_address,
                                  username=my_username, password=my_password)
    dev.show_commands()
    define_inf_exclude(dev)
    dev.parse(dev, bs_dict)
    dev.lag_member_tag(dev)
    dev.delete_info(dev)
    description_bs_parse(dev)
    dev.define_port_bs(dev)
    shorten_bs(dev)
    dev.make_config(dev)
    configure(dev, settings)
    dev.ssh_conn.disconnect()
    dev_queue.task_done()


#######################################################################################
# ------------------------------ main part -------------------------------------------#
#######################################################################################

start_time = datetime.now()
current_date = start_time.strftime("%Y.%m.%d")
current_time = start_time.strftime("%H.%M")

log_folder = Path(f"{Path.cwd()}/logs/{current_date}/")  # current dir / logs / date /
log_folder.mkdir(exist_ok=True)

q = queue.Queue()

settings = get_arguments(argv)
username, password = get_user_pw()
devices = get_device_info("devices.csv")
mac_bs, mac_bs_backup = load_excel(current_date, current_time)  # 04bd.70dc.a7ee : TA7175, информация от МТС

total_devices = len(devices)

print(
    "\n"
    f"Total devices: {total_devices}\n"
    "-------------------------------------------------------------------------------------------------------\n"
    "hostname               ip address      comment\n"
    "---------------------- --------------- ----------------------------------------------------------------\n"
)

for i in range(settings["maxth"]):
    thread = Thread(target=connect_device, args=(username, password, q, mac_bs, mac_bs_backup, settings))
    # thread = Thread(target=test_connect, args=(q, settings, mac_bs, mac_bs_backup))
    # thread = Thread(target=test_connect2, args=(username, password, q, mac_bs, argv_dict))
    thread.daemon = True
    thread.start()

for device in devices:
    q.put(device)

q.join()

failed_connection_count = write_logs(devices, current_time, log_folder, settings)
duration = datetime.now() - start_time
duration_time = timedelta(seconds=duration.seconds)

print("\n"
      "-------------------------------------------------------------------------------------------------------\n"
      f"failed connection:.....{failed_connection_count}\n"
      f"elapsed time:..........{duration_time}\n"
      "-------------------------------------------------------------------------------------------------------")
